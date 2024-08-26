import argparse
from openpyxl import *
from openpyxl.styles import PatternFill
from math import ceil

def ajust_row_height(ws):
    for row in ws.iter_rows():
        line_count = 1
        for cell in row:
            max_char_in_row_width = ws.column_dimensions[cell.column_letter].width
            nb_lines_in_cell = ceil(len(str(cell.value)) / max_char_in_row_width)
            line_count = max(line_count, nb_lines_in_cell)
        ws.row_dimensions[row[0].row].height = line_count * 15

def apply_style(ws):
    dark_fill = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')
    light_fill = PatternFill(start_color='FFF3F3F3', end_color='FFF3F3F3', fill_type='solid')
    count = 0
    for row in ws.iter_rows(min_row=2):
        if count % 2:
            fill = light_fill
        else:
            fill = dark_fill
        for cell in row:
            cell.fill = fill
        count = count + 1

def define_print_area(ws, component_group: int):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_area = 'A1:E' + str(1 + component_group)
    ws.print_title_rows = '1:1'

def fill_with_empty_cells(ws):
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.value == "":
                c.value = " "

def remove_dnf_sheet(wb):
    if 'DNF' in wb.sheetnames:
        del wb['DNF']

def remove_header(ws) -> int:
    component_group = int(ws['B1'].value)
    ws.delete_rows(1, 6)
    c = ws['A2']
    ws.freeze_panes = c 
    return component_group

def sort_components(ws):
    bom = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bom.append(row)
    bom.sort(key=lambda row: row[0])
    ws.delete_rows(2, ws.max_row - 1)
    for row in bom:
        ws.append(row)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input", required=True, type=str)
    parser.add_argument("-o", "--output", type=str)
    args = parser.parse_args()
    if (args.output is None):
        args.output = args.input

    wb = load_workbook(args.input)
    remove_dnf_sheet(wb)
    ws = wb.active
    component_group = remove_header(ws)
    define_print_area(ws, component_group)
    fill_with_empty_cells(ws)
    # sort_components(ws)
    # apply_style(ws)
    ajust_row_height(ws)
    wb.save(args.output)
